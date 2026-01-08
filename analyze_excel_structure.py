import openpyxl
import pandas as pd

# Load the Excel file
file_path = r"data/raw/Occupation des locaux_Automne_2025-2026 (1).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.worksheets[0]

print("=" * 80)
print("EXCEL FILE STRUCTURE ANALYSIS")
print("=" * 80)

# Print header rows
print("\n--- ROW 5 (Group Names) ---")
row5_values = []
for col in range(1, 40):
    val = sheet.cell(row=5, column=col).value
    if val:
        row5_values.append(f"Col {col}: {val}")
print("\n".join(row5_values[:15]))

print("\n--- ROW 6 (Capacities) ---")
row6_values = []
for col in range(1, 40):
    val = sheet.cell(row=6, column=col).value
    if val:
        row6_values.append(f"Col {col}: {val}")
print("\n".join(row6_values[:15]))

print("\n--- ROW 7 (Room Names) ---")
row7_values = []
for col in range(1, 40):
    val = sheet.cell(row=7, column=col).value
    if val:
        row7_values.append(f"Col {col}: {val}")
print("\n".join(row7_values[:15]))

# Sample data cells to understand the structure
print("\n--- SAMPLE DATA ROWS (Rows 8-20) ---")
print("\nLooking for sessions with sub-group indicators...")

# Check for TC S3 sessions specifically
print("\n--- Searching for TC S3 TD Sessions ---")
tc_s3_sessions = []
for row in range(8, min(sheet.max_row + 1, 100)):
    for col in range(3, 40):
        val = sheet.cell(row=row, column=col).value
        if val and "TC S3" in str(val) and "TD" in str(val):
            day = sheet.cell(row=row, column=1).value
            time = sheet.cell(row=row, column=2).value
            room = sheet.cell(row=7, column=col).value
            group = sheet.cell(row=5, column=col).value
            tc_s3_sessions.append({
                'row': row,
                'col': col,
                'session': val,
                'day': day,
                'time': time,
                'room': room,
                'group': group
            })

print(f"\nFound {len(tc_s3_sessions)} TC S3 TD sessions:")
for i, sess in enumerate(tc_s3_sessions[:10]):
    print(f"{i+1}. Row {sess['row']}, Col {sess['col']}: {sess['session']}")
    print(f"   Day: {sess['day']}, Time: {sess['time']}, Room: {sess['room']}, Group Header: {sess['group']}")

# Check for TC S1 sessions
print("\n--- Searching for TC S1 TD Sessions ---")
tc_s1_sessions = []
for row in range(8, min(sheet.max_row + 1, 100)):
    for col in range(3, 40):
        val = sheet.cell(row=row, column=col).value
        if val and "TC S1" in str(val) and "TD" in str(val):
            day = sheet.cell(row=row, column=1).value
            time = sheet.cell(row=row, column=2).value
            room = sheet.cell(row=7, column=col).value
            group = sheet.cell(row=5, column=col).value
            tc_s1_sessions.append({
                'row': row,
                'col': col,
                'session': val,
                'day': day,
                'time': time,
                'room': room,
                'group': group
            })

print(f"\nFound {len(tc_s1_sessions)} TC S1 TD sessions:")
for i, sess in enumerate(tc_s1_sessions[:10]):
    print(f"{i+1}. Row {sess['row']}, Col {sess['col']}: {sess['session']}")
    print(f"   Day: {sess['day']}, Time: {sess['time']}, Room: {sess['room']}, Group Header: {sess['group']}")

# Analysis: Check if session names contain sub-group info
print("\n--- ANALYSIS: Sub-group Identifiers in Session Names ---")
unique_sessions = set()
for row in range(8, min(sheet.max_row + 1, 200)):
    for col in range(3, 40):
        val = sheet.cell(row=row, column=col).value
        if val and isinstance(val, str):
            unique_sessions.add(val.strip())

# Filter TC sessions
tc_sessions = [s for s in unique_sessions if "TC" in s and ("S1" in s or "S3" in s)]
tc_sessions_sorted = sorted(tc_sessions)

print("\nAll unique TC S1/S3 sessions found:")
for sess in tc_sessions_sorted:
    print(f"  - {sess}")

# Check for patterns: GP, GI, GEG, GB
print("\n--- Sub-group Pattern Detection ---")
patterns = ['GP', 'GI', 'GEG', 'GB', 'MTU']
for pattern in patterns:
    matching = [s for s in tc_sessions_sorted if pattern in s]
    print(f"\n{pattern} sessions ({len(matching)}):")
    for sess in matching[:5]:
        print(f"  - {sess}")
