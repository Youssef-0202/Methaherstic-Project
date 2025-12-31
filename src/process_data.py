import pandas as pd
import openpyxl
import os
import re
from datetime import datetime, timedelta

# File paths
INPUT_FILE = r"data/raw/Occupation des locaux_Automne_2025-2026 (1).xlsx"
OUTPUT_DIR = r"data/processed"

def process_data():
    print(f"Processing {INPUT_FILE}...")
    
    # Load Workbook with openpyxl to handle merged cells
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    sheet = wb.worksheets[0] # Assume first sheet
    
    # --- 1. Extract Rooms & Capacities ---
    # Row 7 (index 7, 1-based) -> Room Names (A1, A2...)
    # Row 6 (index 6, 1-based) -> Capacities (400, 400...)
    
    rooms = []
    # Columns 3 to end (C is 3)
    # Iterate columns starting from 3
    col_idx = 3
    room_col_map = {} # Map column index to room_id
    
    while True:
        room_name_cell = sheet.cell(row=7, column=col_idx).value
        if not room_name_cell:
            break # Stop if no more rooms
            
        room_name = str(room_name_cell).strip()
        cap_cell = sheet.cell(row=6, column=col_idx).value
        
        try:
            capacity = int(cap_cell) if cap_cell and isinstance(cap_cell, (int, float)) else 40
        except:
            capacity = 40
            
        # Determine type
        r_type = 'Amphitheater' if room_name.startswith('A') and len(room_name) < 3 else 'Classroom'
        
        rooms.append({
            'room_id': room_name,
            'capacity': capacity,
            'type': r_type
        })
        
        room_col_map[col_idx] = room_name
        col_idx += 1
        
    df_rooms = pd.DataFrame(rooms)
    print(f"Extracted {len(df_rooms)} rooms.")

    # --- 2. Build Time Map ---
    # Map Row Index -> (Day, Time)
    # Rows start from 7 (data starts)
    # Column 1 = Day, Column 2 = Time
    
    row_time_map = {}
    current_day = None
    
    # Scan rows to build time index
    max_row = sheet.max_row
    for r in range(7, max_row + 1):
        day_cell = sheet.cell(row=r, column=1).value
        time_cell = sheet.cell(row=r, column=2).value
        
        if day_cell:
            current_day = str(day_cell).strip()
        
        if time_cell:
            # Format time
            if isinstance(time_cell, (datetime, pd.Timestamp)):
                # Shift back by 30 minutes as requested
                new_time = time_cell - timedelta(minutes=30)
                t_str = new_time.strftime("%H:%M")
            else:
                # Handle string time (e.g. "09:00:00")
                try:
                    t_str = str(time_cell)[:5]
                    t_dt = datetime.strptime(t_str, "%H:%M")
                    new_time = t_dt - timedelta(minutes=30)
                    t_str = new_time.strftime("%H:%M")
                    row_time_map[r] = {'day': current_day, 'time': t_str}
                except:
                    # Invalid time format (e.g. "Matin", "Pause"), skip
                    continue

    # --- 3. Extract Assignments (Handling Merged Cells) ---
    assignments = []
    courses = set()
    processed_cells = set() # Track (row, col) to avoid duplicates if we iterate
    
    # Helper to parse content
    def parse_content(content):
        parts = content.split('-')
        c_id = parts[0].strip()
        g_id = parts[-1].strip() if len(parts) > 1 else "All"
        c_type = 'TD' if 'TD' in content else ('TP' if 'TP' in content else 'Cours')
        return c_id, g_id, c_type

    # A. Handle Merged Ranges
    for merged_range in sheet.merged_cells.ranges:
        # Get bounds
        min_col, min_row, max_col, max_row = merged_range.bounds
        
        # Check if this merge is in the data area
        if min_row >= 7 and min_col >= 3:
            # Get Value from top-left cell
            val = sheet.cell(row=min_row, column=min_col).value
            
            if val and str(val).strip():
                content = str(val).strip()
                
                # Get Room
                if min_col in room_col_map:
                    room_id = room_col_map[min_col]
                    
                    # Get Start Time
                    if min_row in row_time_map:
                        start_info = row_time_map[min_row]
                        day = start_info['day']
                        start_time = start_info['time']
                        
                        # Calculate Duration (slots * 30 mins)
                        # Assuming contiguous rows are 30 min increments
                        slots = (max_row - min_row) + 1
                        duration_hours = (slots * 30) / 60.0
                        
                        # Format as string "4h" or "1.5h"
                        dur_str = f"{int(duration_hours)}h" if duration_hours.is_integer() else f"{duration_hours}h"
                        
                        c_id, g_id, c_type = parse_content(content)
                        
                        assignments.append({
                            'day': day,
                            'start_time': start_time,
                            'duration': dur_str,
                            'room_id': room_id,
                            'course_name': c_id,
                            'group_id': g_id,
                            'type': c_type,
                            'teacher_id': 'Unknown'
                        })
                        courses.add(c_id)
                        
                        # Mark cells as processed
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                processed_cells.add((r, c))

    # B. Handle Single Cells (Non-merged)
    for r in range(7, max_row + 1):
        if r not in row_time_map: continue
        
        for c in room_col_map.keys():
            if (r, c) in processed_cells:
                continue
                
            val = sheet.cell(row=r, column=c).value
            if val and str(val).strip():
                content = str(val).strip()
                
                start_info = row_time_map[r]
                c_id, g_id, c_type = parse_content(content)
                
                assignments.append({
                    'day': start_info['day'],
                    'start_time': start_info['time'],
                    'duration': "0.5h",
                    'room_id': room_col_map[c],
                    'course_name': c_id,
                    'group_id': g_id,
                    'type': c_type,
                    'teacher_id': 'Unknown'
                })
                courses.add(c_id)

    df_assignments = pd.DataFrame(assignments)
    df_courses = pd.DataFrame({'course_name': list(courses)})
    
    # --- SORTING ---
    # Define day order
    day_order = {
        'LUNDI': 0, 'MARDI': 1, 'MERCREDI': 2, 
        'JEUDI': 3, 'VENDREDI': 4, 'SAMEDI': 5, 'DIMANCHE': 6
    }
    
    # Add temporary sort column
    df_assignments['day_index'] = df_assignments['day'].map(lambda x: day_order.get(str(x).upper(), 99))
    
    # Sort by Day Index then Start Time
    df_assignments.sort_values(by=['day_index', 'start_time'], inplace=True)
    
    # Remove temporary column
    df_assignments.drop(columns=['day_index'], inplace=True)
    
    # --- ENRICHMENT (Mocking) ---
    import random
    random.seed(42)
    
    # Set teacher_id to empty (to be filled manually later)
    df_assignments['teacher_id'] = ''
    
    unique_groups = df_assignments['group_id'].unique()
    group_data = [{'group_id': g, 'size': 100 if "All" in g or "TC" in g else 30} for g in unique_groups]
    df_groups = pd.DataFrame(group_data)
    
    # Create empty teachers file (to be filled manually)
    df_teachers = pd.DataFrame({'teacher_id': []})

    print(f"Extracted {len(df_assignments)} assignments.")
    print(f"Extracted {len(df_courses)} unique courses.")
    
    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df_rooms.to_csv(os.path.join(OUTPUT_DIR, 'rooms.csv'), index=False)
    df_assignments.to_csv(os.path.join(OUTPUT_DIR, 'assignments.csv'), index=False)
    df_courses.to_csv(os.path.join(OUTPUT_DIR, 'courses.csv'), index=False)
    df_teachers.to_csv(os.path.join(OUTPUT_DIR, 'teachers.csv'), index=False)
    df_groups.to_csv(os.path.join(OUTPUT_DIR, 'groups.csv'), index=False)
    
    print(f"Saved CSVs to {OUTPUT_DIR}")

if __name__ == "__main__":
    process_data()
