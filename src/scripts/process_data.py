import pandas as pd
import openpyxl
import os
import re
from datetime import datetime, timedelta

# File paths
INPUT_FILE = r"data/raw/Occupation des locaux_Automne_2025-2026 (1).xlsx"
OUTPUT_DIR = r"data/processed"

def process_data():
    print(f"Processing {INPUT_FILE}...")
    
    # Load Workbook with openpyxl to handle merged cells
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    sheet = wb.worksheets[0] # Assume first sheet
    
    # --- 1. Extract Rooms & Capacities ---
    # Row 7 (index 7, 1-based) -> Room Names (A1, A2...)
    # Row 6 (index 6, 1-based) -> Capacities (400, 400...)
    
    rooms = []
    # Columns 3 to end (C is 3)
    # Iterate columns starting from 3
    col_idx = 3
    room_col_map = {} # Map column index to room_id
    
    while True:
        room_name_cell = sheet.cell(row=7, column=col_idx).value
        if not room_name_cell:
            break # Stop if no more rooms
            
        room_name = str(room_name_cell).strip()
        cap_cell = sheet.cell(row=6, column=col_idx).value
        
        try:
            capacity = int(cap_cell) if cap_cell and isinstance(cap_cell, (int, float)) else 40
        except:
            capacity = 40
            
        # Determine type
        r_type = 'Amphitheater' if room_name.startswith('A') and len(room_name) < 3 else 'Classroom'
        
        rooms.append({
            'room_id': room_name,
            'capacity': capacity,
            'type': r_type
        })
        
        room_col_map[col_idx] = room_name
        col_idx += 1
        
    df_rooms = pd.DataFrame(rooms)
    print(f"Extracted {len(df_rooms)} rooms.")

    # --- 2. Build Time Map ---
    # Map Row Index -> (Day, Time)
    # Rows start from 8 (data starts)
    # Column 1 = Day, Column 2 = Time
    
    row_time_map = {}
    current_day = None
    
    # Scan rows to build time index
    max_row = sheet.max_row
    for r in range(8, max_row + 1):
        day_cell = sheet.cell(row=r, column=1).value
        time_cell = sheet.cell(row=r, column=2).value
        
        if day_cell:
            current_day = str(day_cell).strip()
        
        if time_cell:
            # Format time
            if isinstance(time_cell, (datetime, pd.Timestamp)):
                # Shift back by 30 minutes as requested
                new_time = time_cell - timedelta(minutes=30)
                t_str = new_time.strftime("%H:%M")
                row_time_map[r] = {'day': current_day, 'time': t_str}
            else:
                # Handle string time (e.g. "09:00:00")
                try:
                    t_str = str(time_cell)[:5]
                    t_dt = datetime.strptime(t_str, "%H:%M")
                    new_time = t_dt - timedelta(minutes=30)
                    t_str = new_time.strftime("%H:%M")
                    row_time_map[r] = {'day': current_day, 'time': t_str}
                except:
                    # Invalid time format (e.g. "Matin", "Pause"), skip
                    continue

    # --- 3. Extract Group Names from Column Headers ---
    # Row 5 has the group names (ERME, IFA, MDSIM, GE, MEA, etc.)
    group_col_map = {}
    for col_idx in room_col_map.keys():
        group_cell = sheet.cell(row=5, column=col_idx).value
        if group_cell:
            group_name = str(group_cell).strip()
            group_col_map[col_idx] = group_name
        else:
            group_col_map[col_idx] = "Unknown"
    
    # --- 4. Extract Assignments (Handling Merged Cells) ---
    assignments = []
    sessions = set()
    processed_cells = set()
    
    # Helper to parse content
    def parse_content(content):
        # Content is the session info (e.g., "TC S3 - TD")
        session_type = 'TD' if 'TD' in content else ('TP' if 'TP' in content else 'Cours')
        session_name = content.strip()
        return session_name, session_type

    # A. Handle Merged Ranges
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        
        # Check if this merge is in the data area (starts from row 8)
        if min_row >= 8 and min_col >= 3:
            val = sheet.cell(row=min_row, column=min_col).value
            
            if val and str(val).strip():
                content = str(val).strip()
                
                # Get Room (merged cells usually stay in one column for room, or span multiple)
                # We assume the room is associated with the columns it spans
                if min_col in room_col_map:
                    room_id = room_col_map[min_col]
                    
                    if min_row in row_time_map:
                        start_info = row_time_map[min_row]
                        day = start_info['day']
                        start_time = start_info['time']
                        
                        slots = (max_row - min_row) + 1
                        duration_hours = (slots * 30) / 60.0
                        dur_str = f"{int(duration_hours)}h" if duration_hours.is_integer() else f"{duration_hours}h"
                        
                        session_name, session_type = parse_content(content)
                        
                        # Collect ALL groups involved in this merged range
                        involved_groups = []
                        for c in range(min_col, max_col + 1):
                            g_name = group_col_map.get(c, "Unknown")
                            if g_name not in involved_groups:
                                involved_groups.append(g_name)
                        
                        groups_str = ", ".join(involved_groups)
                        
                        assignments.append({
                            'day': day,
                            'start_time': start_time,
                            'duration': dur_str,
                            'room_id': room_id,
                            'session_name': session_name,
                            'session_type': session_type,
                            'teacher_id': ''
                        })
                        sessions.add(session_name)
                        
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                processed_cells.add((r, c))

    # B. Handle Single Cells (Non-merged)
    for r in range(8, max_row + 1):
        if r not in row_time_map: continue
        
        for c in room_col_map.keys():
            if (r, c) in processed_cells:
                continue
                
            val = sheet.cell(row=r, column=c).value
            if val and str(val).strip():
                content = str(val).strip()
                
                start_info = row_time_map[r]
                session_name, session_type = parse_content(content)
                
                assignments.append({
                    'day': start_info['day'],
                    'start_time': start_info['time'],
                    'duration': "0.5h",
                    'room_id': room_col_map[c],
                    'session_name': session_name,
                    'session_type': session_type,
                    'teacher_id': ''
                })
                sessions.add(session_name)

    df_assignments = pd.DataFrame(assignments)
    df_sessions = pd.DataFrame({'session_name': list(sessions)})
    
    # --- SORTING ---
    day_order = {'LUNDI': 0, 'MARDI': 1, 'MERCREDI': 2, 'JEUDI': 3, 'VENDREDI': 4, 'SAMEDI': 5, 'DIMANCHE': 6}
    df_assignments['day_index'] = df_assignments['day'].map(lambda x: day_order.get(str(x).upper(), 99))
    df_assignments.sort_values(by=['day_index', 'start_time'], inplace=True)
    df_assignments.drop(columns=['day_index'], inplace=True)
    
    print(f"Extracted {len(df_assignments)} assignments.")
    print(f"Extracted {len(df_sessions)} unique sessions.")
    
    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df_rooms.to_csv(os.path.join(OUTPUT_DIR, 'rooms.csv'), index=False)
    df_assignments.to_csv(os.path.join(OUTPUT_DIR, 'assignments.csv'), index=False)
    df_sessions.to_csv(os.path.join(OUTPUT_DIR, 'sessions.csv'), index=False)
    
    print(f"Saved CSVs to {OUTPUT_DIR}")

if __name__ == "__main__":
    process_data()
