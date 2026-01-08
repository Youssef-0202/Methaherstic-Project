import pandas as pd
import openpyxl
import os
import re
from datetime import datetime, timedelta

# File paths
INPUT_FILE = r"data/raw/Occupation des locaux_Automne_2025-2026 (1).xlsx"
OUTPUT_DIR = r"data/processed"

def extract_subgroup_from_session(session_name, group_header):
    """
    Extract sub-group information from session name.

    Examples:
    - "TC S1 - TD GP & GI" -> ["S1_GP", "S1_GI"]
    - "TC S1 - TD GB & GEG" -> ["S1_GB", "S1_GEG"]
    - "TC S3 - TD" with group_header "Unknown" -> ["S3"]
    - "TC S1" -> ["S1"]
    """
    session_name = session_name.strip()

    # Extract the main group (S1, S2, S3, etc.)
    main_group_match = re.search(r'\b(S[123]|S\d+)\b', session_name)
    if not main_group_match:
        # No S1/S2/S3 found, try other patterns
        if group_header and group_header != "Unknown":
            return [group_header]
        return ["Unknown"]

    main_group = main_group_match.group(1)

    # Look for sub-group indicators
    subgroups = []

    # Pattern 1: "GP & GI", "GB & GEG", etc.
    subgroup_pattern = re.findall(r'\b(GP|GI|GB|GEG|MTU)\b', session_name)

    if subgroup_pattern:
        # Found specific sub-groups
        for sg in subgroup_pattern:
            subgroups.append(f"{main_group}_{sg}")
    else:
        # No specific sub-group found, return main group
        subgroups.append(main_group)

    return subgroups

def assign_teacher_by_heuristic(session_name, session_type, room_id, teacher_counter):
    """
    Assign teachers based on heuristic rules.
    For TD/TP sessions with same course but different rooms, assign different teachers.
    """
    # Create a unique key for this session context
    course_key = session_name.replace(" - TD", "").replace(" - TP", "").strip()

    if session_type in ["TD", "TP"]:
        # Different rooms for same TD/TP course should have different teachers
        key = f"{course_key}_{session_type}_{room_id}"
    else:
        # Cours sessions: same course always same teacher
        key = f"{course_key}_{session_type}"

    if key not in teacher_counter:
        teacher_counter[key] = f"T_{len(teacher_counter) + 1:03d}"

    return teacher_counter[key]

def process_data_with_subgroups():
    print(f"Processing {INPUT_FILE} with sub-group extraction...")

    # Load Workbook
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    sheet = wb.worksheets[0]

    # --- 1. Extract Rooms & Capacities ---
    rooms = []
    col_idx = 3
    room_col_map = {}

    while True:
        room_name_cell = sheet.cell(row=7, column=col_idx).value
        if not room_name_cell:
            break

        room_name = str(room_name_cell).strip()
        cap_cell = sheet.cell(row=6, column=col_idx).value

        try:
            capacity = int(cap_cell) if cap_cell and isinstance(cap_cell, (int, float)) else 48
        except:
            capacity = 48

        r_type = 'Amphitheater' if room_name.startswith('A') and len(room_name) <= 2 else 'Classroom'

        rooms.append({
            'room_id': room_name,
            'capacity': capacity,
            'type': r_type
        })

        room_col_map[col_idx] = room_name
        col_idx += 1

    df_rooms = pd.DataFrame(rooms)
    print(f"Extracted {len(df_rooms)} rooms.")

    # --- 2. Build Time Map ---
    row_time_map = {}
    current_day = None

    max_row = sheet.max_row
    for r in range(8, max_row + 1):
        day_cell = sheet.cell(row=r, column=1).value
        time_cell = sheet.cell(row=r, column=2).value

        if day_cell:
            current_day = str(day_cell).strip().upper()

        if time_cell:
            if isinstance(time_cell, (datetime, pd.Timestamp)):
                new_time = time_cell - timedelta(minutes=30)
                t_str = new_time.strftime("%H:%M")
                row_time_map[r] = {'day': current_day, 'time': t_str}
            else:
                try:
                    t_str = str(time_cell)[:5]
                    t_dt = datetime.strptime(t_str, "%H:%M")
                    new_time = t_dt - timedelta(minutes=30)
                    t_str = new_time.strftime("%H:%M")
                    row_time_map[r] = {'day': current_day, 'time': t_str}
                except:
                    continue

    # --- 3. Extract Group Names from Column Headers ---
    group_col_map = {}
    for col_idx in room_col_map.keys():
        group_cell = sheet.cell(row=5, column=col_idx).value
        if group_cell:
            group_name = str(group_cell).strip()
            group_col_map[col_idx] = group_name
        else:
            group_col_map[col_idx] = "Unknown"

    # --- 4. Extract Assignments with Sub-group Detection ---
    assignments = []
    sessions_set = set()
    processed_cells = set()
    teacher_counter = {}

    def parse_content(content):
        session_type = 'TD' if 'TD' in content else ('TP' if 'TP' in content else 'Cours')
        session_name = content.strip()
        return session_name, session_type

    # A. Handle Merged Ranges
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        if min_row >= 8 and min_col >= 3:
            val = sheet.cell(row=min_row, column=min_col).value

            if val and str(val).strip():
                content = str(val).strip()

                if min_col in room_col_map:
                    room_id = room_col_map[min_col]

                    if min_row in row_time_map:
                        start_info = row_time_map[min_row]
                        day = start_info['day']
                        start_time = start_info['time']

                        slots = (max_row - min_row) + 1
                        duration_hours = (slots * 30) / 60.0
                        dur_str = f"{int(duration_hours)}h" if duration_hours.is_integer() else f"{duration_hours}h"

                        session_name, session_type = parse_content(content)

                        # Get group header for this column
                        group_header = group_col_map.get(min_col, "Unknown")

                        # Extract sub-groups from session name
                        subgroups = extract_subgroup_from_session(session_name, group_header)
                        groups_str = ";".join(subgroups)

                        # Assign teacher
                        teacher_id = assign_teacher_by_heuristic(session_name, session_type, room_id, teacher_counter)

                        assignments.append({
                            'day': day,
                            'start_time': start_time,
                            'duration': dur_str,
                            'room_id': room_id,
                            'session_name': session_name,
                            'session_type': session_type,
                            'teacher_id': teacher_id,
                            'involved_groups': groups_str
                        })
                        sessions_set.add(session_name)

                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                processed_cells.add((r, c))

    # B. Handle Single Cells (Non-merged)
    for r in range(8, max_row + 1):
        if r not in row_time_map:
            continue

        for c in room_col_map.keys():
            if (r, c) in processed_cells:
                continue

            val = sheet.cell(row=r, column=c).value
            if val and str(val).strip():
                content = str(val).strip()

                start_info = row_time_map[r]
                session_name, session_type = parse_content(content)

                group_header = group_col_map.get(c, "Unknown")
                subgroups = extract_subgroup_from_session(session_name, group_header)
                groups_str = ";".join(subgroups)

                teacher_id = assign_teacher_by_heuristic(session_name, session_type, room_col_map[c], teacher_counter)

                assignments.append({
                    'day': start_info['day'],
                    'start_time': start_info['time'],
                    'duration': "0.5h",
                    'room_id': room_col_map[c],
                    'session_name': session_name,
                    'session_type': session_type,
                    'teacher_id': teacher_id,
                    'involved_groups': groups_str
                })
                sessions_set.add(session_name)

    df_assignments = pd.DataFrame(assignments)

    # --- SORTING ---
    day_order = {'LUNDI': 0, 'MARDI': 1, 'MERCREDI': 2, 'JEUDI': 3, 'VENDREDI': 4, 'SAMEDI': 5, 'DIMANCHE': 6}
    df_assignments['day_index'] = df_assignments['day'].map(lambda x: day_order.get(str(x).upper(), 99))
    df_assignments.sort_values(by=['day_index', 'start_time'], inplace=True)
    df_assignments.drop(columns=['day_index'], inplace=True)

    print(f"Extracted {len(df_assignments)} assignments.")
    print(f"Extracted {len(sessions_set)} unique sessions.")
    print(f"Created {len(teacher_counter)} unique teacher assignments.")

    # --- 5. Create Teachers CSV ---
    teachers = []
    for teacher_id in sorted(teacher_counter.values(), key=lambda x: int(x.split('_')[1])):
        # Try to infer specialization from assigned courses
        assigned_sessions = df_assignments[df_assignments['teacher_id'] == teacher_id]['session_name'].values
        if len(assigned_sessions) > 0:
            first_session = assigned_sessions[0]
            # Extract specialization hint
            if 'MST' in first_session:
                spec = 'MST'
            elif 'TC' in first_session:
                spec = 'TC'
            elif 'IEEA' in first_session:
                spec = 'IEEA'
            elif 'GC' in first_session:
                spec = 'GC'
            else:
                spec = 'General'
        else:
            spec = 'General'

        teachers.append({
            'teacher_id': teacher_id,
            'name': f'Professor_{teacher_id}',
            'specialization': spec
        })

    df_teachers = pd.DataFrame(teachers)
    print(f"Created {len(df_teachers)} teacher records.")

    # --- 6. Create Enhanced Groups CSV with Sub-groups ---
    groups = []
    all_groups_mentioned = set()

    for assignment in assignments:
        groups_list = assignment['involved_groups'].split(';')
        for g in groups_list:
            all_groups_mentioned.add(g.strip())

    # Categorize groups
    for group_name in sorted(all_groups_mentioned):
        if group_name == "Unknown":
            continue

        # Determine section and size
        if group_name.startswith('S1') or group_name.startswith('S2') or group_name.startswith('S3'):
            section = 'TC'
            # Sub-groups have smaller size
            if '_' in group_name:
                size = 45  # Sub-group size
            else:
                size = 180  # Main group size
        elif group_name in ['MDSIM', 'GE', 'MEA', 'RD.EVG', 'RD', 'PNB']:
            section = 'MST'
            size = 30
        elif group_name in ['ISA', 'GMP', 'ERME', 'IFA', 'PCM', 'IEGS']:
            section = 'CI'
            size = 45
        elif group_name in ['GC', 'IEEA', 'EE', 'GARM', 'MIASI', 'BP', 'BA', 'SIR']:
            section = 'LST'
            size = 70
        else:
            section = 'Unknown'
            size = 40

        groups.append({
            'group_name': group_name,
            'section': section,
            'size': size
        })

    df_groups = pd.DataFrame(groups)
    print(f"Created {len(df_groups)} group definitions (including sub-groups).")

    # --- 7. Save all CSVs ---
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df_rooms.to_csv(os.path.join(OUTPUT_DIR, 'rooms.csv'), index=False)
    df_assignments.to_csv(os.path.join(OUTPUT_DIR, 'assignments.csv'), index=False)
    df_teachers.to_csv(os.path.join(OUTPUT_DIR, 'teachers.csv'), index=False)
    df_groups.to_csv(os.path.join(OUTPUT_DIR, 'groups.csv'), index=False)

    print(f"\n{'='*60}")
    print(f"Successfully saved all CSVs to {OUTPUT_DIR}")
    print(f"{'='*60}")

    # Print summary statistics
    print("\n--- SUMMARY ---")
    print(f"Total assignments: {len(df_assignments)}")
    print(f"Assignments with Unknown groups: {len(df_assignments[df_assignments['involved_groups'].str.contains('Unknown')])}")
    print(f"Total groups (including sub-groups): {len(df_groups)}")
    print(f"Total teachers: {len(df_teachers)}")
    print(f"Total rooms: {len(df_rooms)}")

    # Check for improvements
    print("\n--- SUB-GROUP DETECTION RESULTS ---")
    subgroup_assignments = df_assignments[df_assignments['involved_groups'].str.contains('_')]
    print(f"Assignments with detected sub-groups: {len(subgroup_assignments)}")

    print("\nSample sub-group assignments:")
    for idx, row in subgroup_assignments.head(10).iterrows():
        print(f"  {row['session_name']} -> Groups: {row['involved_groups']}")

if __name__ == "__main__":
    process_data_with_subgroups()
